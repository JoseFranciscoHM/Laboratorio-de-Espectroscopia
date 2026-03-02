import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import serial
import serial.tools.list_ports
import time
import threading
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import Workbook
import numpy as np
from datetime import datetime
import json
import os

class FotoluminiscenciaApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Sistema de Medición de Fotoluminiscencia")
        self.geometry("1000x700")
        self.configure(bg="#f0f0f0")
        
        # Configurar estilo
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.configure_styles()
        
        self.data = []  # Inicializar lista de datos
        self.measuring = False  # Bandera para controlar la medición
        self.mono_ser = None
        self.lockin_ser = None
        
        # Cargar configuración previa
        self.load_config()
        
        self.create_widgets()
        
        # Proteger contra cierre durante medición
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def configure_styles(self):
        # Configurar estilos para los widgets
        self.style.configure('TFrame', background="#f0f0f0")
        self.style.configure('TLabel', background="#f0f0f0", foreground="#333333", font=('Arial', 9))
        self.style.configure('TButton', font=('Arial', 9), background="#4CAF50", foreground="white")
        self.style.configure('TEntry', font=('Arial', 9), fieldbackground="white", foreground="#333333")
        self.style.configure('TCombobox', font=('Arial', 9), fieldbackground="white", foreground="#333333")
        self.style.configure('TLabelframe', background="#f0f0f0", foreground="#2c3e50")
        self.style.configure('TLabelframe.Label', background="#f0f0f0", foreground="#2c3e50")
        
        self.style.map('TButton', 
                      background=[('active', '#45a049'), ('disabled', '#cccccc')],
                      foreground=[('active', 'white'), ('disabled', '#888888')])

    def load_config(self):
        """Cargar configuración previa si existe"""
        self.config_data = {
            "mono_port": "COM4",
            "lockin_port": "COM5",
            "start_wl": "400",
            "end_wl": "700",
            "step": "1",
            "readings": "3",
            "wait_time": "0.5"
        }
        
        try:
            if os.path.exists("config_pl.json"):
                with open("config_pl.json", "r") as f:
                    self.config_data = json.load(f)
        except:
            pass  # Si hay error, se usan los valores por defecto

    def save_config(self):
        """Guardar configuración actual"""
        try:
            with open("config_pl.json", "w") as f:
                json.dump(self.config_data, f)
        except:
            pass  # Si no se puede guardar, continuar

    def create_widgets(self):
        # Frame de configuración
        config_frame = ttk.LabelFrame(self, text="Configuración del Experimento", padding="10")
        config_frame.pack(fill="x", padx=10, pady=5)

        # Puerto del monocromador
        ttk.Label(config_frame, text="Puerto Monocromador:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        
        mono_frame = self.style.configure('TLabelframe', background="#f0f0f0", foreground="#2c3e50")
        self.style.configure('TLabelframe.Label', background="#f0f0f0", foreground="#2c3e50")
        mono_frame.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        
        self.monochromator_port = ttk.Combobox(mono_frame, width=15, values=self.get_serial_ports())
        self.monochromator_port.set(self.config_data["mono_port"])
        self.monochromator_port.pack(side=tk.LEFT)
        
        ttk.Button(mono_frame, text="Refrescar", command=self.refresh_ports, width=8).pack(side=tk.LEFT, padx=(5, 0))

        # Puerto del lock-in amplificador
        ttk.Label(config_frame, text="Puerto Lock-in:").grid(row=0, column=2, padx=5, pady=5, sticky="e")
        
        lockin_frame = ttk.Frame(config_frame)
        lockin_frame.grid(row=0, column=3, padx=5, pady=5, sticky="ew")
        
        self.lockin_port = ttk.Combobox(lockin_frame, width=15, values=self.get_serial_ports())
        self.lockin_port.set(self.config_data["lockin_port"])
        self.lockin_port.pack(side=tk.LEFT)
        
        ttk.Button(lockin_frame, text="Test", command=self.test_ports, width=8).pack(side=tk.LEFT, padx=(5, 0))

        # Longitud de onda inicial
        ttk.Label(config_frame, text="Longitud de onda inicial (nm):").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.start_wavelength = ttk.Entry(config_frame, width=15)
        self.start_wavelength.insert(0, self.config_data["start_wl"])
        self.start_wavelength.grid(row=1, column=1, padx=5, pady=5)

        # Longitud de onda final
        ttk.Label(config_frame, text="Longitud de onda final (nm):").grid(row=1, column=2, padx=5, pady=5, sticky="e")
        self.end_wavelength = ttk.Entry(config_frame, width=15)
        self.end_wavelength.insert(0, self.config_data["end_wl"])
        self.end_wavelength.grid(row=1, column=3, padx=5, pady=5)

        # Paso de longitud de onda
        ttk.Label(config_frame, text="Paso (nm):").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.step = ttk.Entry(config_frame, width=15)
        self.step.insert(0, self.config_data["step"])
        self.step.grid(row=2, column=1, padx=5, pady=5)

        # Número de lecturas por longitud de onda
        ttk.Label(config_frame, text="Lecturas por punto:").grid(row=2, column=2, padx=5, pady=5, sticky="e")
        self.readings_per_wavelength = ttk.Entry(config_frame, width=15)
        self.readings_per_wavelength.insert(0, self.config_data["readings"])
        self.readings_per_wavelength.grid(row=2, column=3, padx=5, pady=5)

        # Tiempo de espera entre movimientos
        ttk.Label(config_frame, text="Tiempo de espera (s):").grid(row=3, column=0, padx=5, pady=5, sticky="e")
        self.wait_time = ttk.Entry(config_frame, width=15)
        self.wait_time.insert(0, self.config_data["wait_time"])
        self.wait_time.grid(row=3, column=1, padx=5, pady=5)

        # Botones de control
        button_frame = ttk.Frame(config_frame)
        button_frame.grid(row=3, column=2, columnspan=2, padx=5, pady=5, sticky="e")

        self.start_button = ttk.Button(button_frame, text="Iniciar Medición", command=self.start_measurement)
        self.start_button.pack(side=tk.LEFT, padx=5)

        self.stop_button = ttk.Button(button_frame, text="Detener", command=self.stop_measurement, state=tk.DISABLED)
        self.stop_button.pack(side=tk.LEFT, padx=5)

        self.save_button = ttk.Button(button_frame, text="Guardar Datos", command=self.save_to_excel)
        self.save_button.pack(side=tk.LEFT, padx=5)

        # Frame para la gráfica
        graph_frame = ttk.LabelFrame(self, text="Gráfica en Tiempo Real", padding="10")
        graph_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.figure, self.ax = plt.subplots(figsize=(8, 5))
        self.ax.set_xlabel("Longitud de onda (nm)", fontsize=10)
        self.ax.set_ylabel("Intensidad (V)", fontsize=10)
        self.ax.grid(True, alpha=0.3)
        self.canvas = FigureCanvasTkAgg(self.figure, master=graph_frame)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        
        # Toolbar para la gráfica
        toolbar_frame = ttk.Frame(graph_frame)
        toolbar_frame.pack(fill="x", pady=(0, 5))
        
        ttk.Button(toolbar_frame, text="Zoom +", command=self.zoom_in).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar_frame, text="Zoom -", command=self.zoom_out).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar_frame, text="Autoajustar", command=self.autoscale_plot).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar_frame, text="Limpiar Gráfica", command=self.clear_plot).pack(side=tk.LEFT, padx=2)

        # Frame de estado
        status_frame = ttk.Frame(self)
        status_frame.pack(fill="x", padx=10, pady=5)
        
        self.status_var = tk.StringVar(value="Listo")
        status_label = ttk.Label(status_frame, textvariable=self.status_var, foreground="#2c3e50")
        status_label.pack(side=tk.LEFT)

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(status_frame, variable=self.progress_var, mode='determinate')
        self.progress_bar.pack(side=tk.RIGHT, fill="x", expand=True, padx=(10, 0))
        
        # Actualizar lista de puertos
        self.after(100, self.refresh_ports)

    def get_serial_ports(self):
        """Obtener lista de puertos seriales disponibles"""
        ports = serial.tools.list_ports.comports()
        return [port.device for port in ports]

    def refresh_ports(self):
        """Actualizar lista de puertos en los combobox"""
        ports = self.get_serial_ports()
        self.monochromator_port['values'] = ports
        self.lockin_port['values'] = ports

    def test_ports(self):
        """Probar la conexión con los puertos seleccionados"""
        mono_port = self.monochromator_port.get().strip()
        lockin_port = self.lockin_port.get().strip()
        
        success = True
        message = ""
        
        # Probar monocromador
        try:
            with serial.Serial(mono_port, baudrate=9600, timeout=2) as ser:
                ser.write(b"*IDN?\r\n")
                time.sleep(0.5)
                response = ser.readline().decode().strip()
                message += f"Monocromador: {response if response else 'Respuesta recibida'}\n"
        except Exception as e:
            message += f"Error monocromador: {str(e)}\n"
            success = False
            
        # Probar lock-in
        try:
            with serial.Serial(lockin_port, baudrate=9600, timeout=2) as ser:
                ser.write(b"*IDN?\r\n")
                time.sleep(0.5)
                response = ser.readline().decode().strip()
                message += f"Lock-in: {response if response else 'Respuesta recibida'}\n"
        except Exception as e:
            message += f"Error lock-in: {str(e)}\n"
            success = False
            
        if success:
            messagebox.showinfo("Prueba de Puertos", f"Conexión exitosa:\n{message}")
        else:
            messagebox.showerror("Prueba de Puertos", f"Error en la conexión:\n{message}")

    def zoom_in(self):
        """Aumentar zoom de la gráfica"""
        xlim = self.ax.get_xlim()
        ylim = self.ax.get_ylim()
        x_center = np.mean(xlim)
        y_center = np.mean(ylim)
        x_range = (xlim[1] - xlim[0]) * 0.7  # Reducir rango en 30%
        y_range = (ylim[1] - ylim[0]) * 0.7
        
        self.ax.set_xlim(x_center - x_range/2, x_center + x_range/2)
        self.ax.set_ylim(y_center - y_range/2, y_center + y_range/2)
        self.canvas.draw()

    def zoom_out(self):
        """Disminuir zoom de la gráfica"""
        xlim = self.ax.get_xlim()
        ylim = self.ax.get_ylim()
        x_center = np.mean(xlim)
        y_center = np.mean(ylim)
        x_range = (xlim[1] - xlim[0]) * 1.3  # Aumentar rango en 30%
        y_range = (ylim[1] - ylim[0]) * 1.3
        
        self.ax.set_xlim(x_center - x_range/2, x_center + x_range/2)
        self.ax.set_ylim(y_center - y_range/2, y_center + y_range/2)
        self.canvas.draw()

    def autoscale_plot(self):
        """Autoajustar la gráfica"""
        self.ax.relim()
        self.ax.autoscale_view()
        self.canvas.draw()

    def clear_plot(self):
        """Limpiar la gráfica"""
        self.ax.clear()
        self.ax.set_xlabel("Longitud de onda (nm)")
        self.ax.set_ylabel("Intensidad (V)")
        self.ax.grid(True, alpha=0.3)
        self.canvas.draw()

    def start_measurement(self):
        # Guardar configuración actual
        self.config_data = {
            "mono_port": self.monochromator_port.get(),
            "lockin_port": self.lockin_port.get(),
            "start_wl": self.start_wavelength.get(),
            "end_wl": self.end_wavelength.get(),
            "step": self.step.get(),
            "readings": self.readings_per_wavelength.get(),
            "wait_time": self.wait_time.get()
        }
        self.save_config()
        
        try:
            start_wl = float(self.start_wavelength.get())
            end_wl = float(self.end_wavelength.get())
            step = float(self.step.get())
            readings = int(self.readings_per_wavelength.get())
            wait_time = float(self.wait_time.get())
            mono_port = self.monochromator_port.get().strip()
            lockin_port = self.lockin_port.get().strip()
            
            if start_wl >= end_wl:
                messagebox.showerror("Error", "La longitud de onda inicial debe ser menor que la final.")
                return
                
            if step <= 0:
                messagebox.showerror("Error", "El paso debe ser mayor que cero.")
                return

        except ValueError as e:
            messagebox.showerror("Error", f"Por favor, ingrese valores numéricos válidos.\n{str(e)}")
            return

        # Limpiar datos anteriores
        self.data = []
        self.ax.clear()
        self.ax.set_xlabel("Longitud de onda (nm)")
        self.ax.set_ylabel("Intensidad (V)")
        self.ax.grid(True, alpha=0.3)
        
        # Cambiar estado de los botones
        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.save_button.config(state=tk.DISABLED)
        self.measuring = True
        
        # Calcular número total de puntos para la barra de progreso
        total_points = int((end_wl - start_wl) / step) + 1
        self.progress_var.set(0)
        self.progress_bar.config(maximum=total_points)
        
        # Iniciar medición en hilo separado
        threading.Thread(
            target=self.measure, 
            args=(start_wl, end_wl, step, readings, wait_time, mono_port, lockin_port), 
            daemon=True
        ).start()

    def stop_measurement(self):
        self.measuring = False
        self.status_var.set("Medición detenida")
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        self.save_button.config(state=tk.NORMAL)
        
        # Cerrar conexiones seriales si están abiertas
        try:
            if self.mono_ser and self.mono_ser.is_open:
                self.mono_ser.close()
            if self.lockin_ser and self.lockin_ser.is_open:
                self.lockin_ser.close()
        except:
            pass

    def measure(self, start_wl, end_wl, step, readings, wait_time, mono_port, lockin_port):
        try:
            # Abrir conexiones seriales
            self.mono_ser = serial.Serial(mono_port, baudrate=9600, timeout=2)
            self.lockin_ser = serial.Serial(lockin_port, baudrate=9600, timeout=1)
            
            # Configurar lock-in si es necesario
            # self.lockin_ser.write(b"XYZ\r\n")  # Comandos de configuración
            
            # Generar la secuencia de longitudes de onda
            wavelengths = []
            current_wl = start_wl
            while current_wl <= end_wl and self.measuring:
                wavelengths.append(current_wl)
                current_wl += step
            
            intensities = []
            std_devs = []  # Desviaciones estándar
            
            for i, wl in enumerate(wavelengths):
                if not self.measuring:
                    break
                    
                # Actualizar estado
                self.status_var.set(f"Moviendo a {wl:.1f} nm")
                self.progress_var.set(i)
                self.update_idletasks()
                
                # Mover monocromador
                command = f"{wl} GOTO\r\n"
                self.mono_ser.write(command.encode('utf-8'))
                time.sleep(wait_time)  # Esperar a que se mueva
                
                # Leer múltiples veces y promediar
                readings_list = []
                for _ in range(readings):
                    if not self.measuring:
                        break
                        
                    self.lockin_ser.reset_input_buffer()
                    self.lockin_ser.write(b'Q1\r')
                    time.sleep(0.1)
                    response = self.lockin_ser.readline().decode().strip()
                    try:
                        value = float(response)
                        readings_list.append(value)
                    except ValueError:
                        continue
                    time.sleep(0.1)
                
                if readings_list:
                    avg_intensity = sum(readings_list) / len(readings_list)
                    std_dev = np.std(readings_list) if len(readings_list) > 1 else 0
                else:
                    avg_intensity = 0
                    std_dev = 0
                
                intensities.append(avg_intensity)
                std_devs.append(std_dev)
                self.data.append((wl, avg_intensity, std_dev))
                
                # Actualizar gráfica
                self.ax.clear()
                self.ax.errorbar(wavelengths[:len(intensities)], intensities, yerr=std_devs, 
                                fmt='b-', linewidth=1.5, marker='o', markersize=3, capsize=2)
                self.ax.set_xlabel("Longitud de onda (nm)")
                self.ax.set_ylabel("Intensidad (V)")
                self.ax.grid(True, alpha=0.3)
                self.canvas.draw()
            
            if self.measuring:
                self.status_var.set("Medición completada")
                self.save_to_excel()
                messagebox.showinfo("Completado", "La medición ha finalizado y los datos se han guardado.")
            
        except serial.SerialException as e:
            messagebox.showerror("Error de conexión", f"No se pudo conectar a los dispositivos:\n{e}")
            self.status_var.set("Error de conexión")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error inesperado:\n{e}")
            self.status_var.set("Error")
        finally:
            self.measuring = False
            self.start_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)
            self.save_button.config(state=tk.NORMAL)
            
            # Cerrar conexiones seriales
            try:
                if self.mono_ser and self.mono_ser.is_open:
                    self.mono_ser.close()
                if self.lockin_ser and self.lockin_ser.is_open:
                    self.lockin_ser.close()
            except:
                pass

    def save_to_excel(self):
        if not self.data:
            messagebox.showwarning("Sin datos", "No hay datos para guardar.")
            return
            
        try:
            # Pedir al usuario dónde guardar el archivo
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"datos_fotoluminiscencia_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not filename:  # Usuario canceló
                return
                
            wb = Workbook()
            ws = wb.active
            ws.title = "Datos de Fotoluminiscencia"
            ws.append(["Longitud de onda (nm)", "Intensidad (V)", "Desviación estándar"])
            
            for wl, intensity, std_dev in self.data:
                ws.append([wl, intensity, std_dev])
            
            # Añadir gráfico al Excel
            from openpyxl.chart import ScatterChart, Reference, Series
            chart = ScatterChart()
            chart.title = "Espectro de Fotoluminiscencia"
            chart.x_axis.title = "Longitud de onda (nm)"
            chart.y_axis.title = "Intensidad (V)"
            
            x_data = Reference(ws, min_col=1, min_row=2, max_row=len(self.data) + 1)
            y_data = Reference(ws, min_col=2, min_row=2, max_row=len(self.data) + 1)
            
            series = Series(y_data, x_data, title="Intensidad")
            chart.series.append(series)
            
            ws.add_chart(chart, "E2")
            
            # Añadir metadatos del experimento
            ws['A10'] = "Configuración del experimento:"
            ws['A11'] = f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A12'] = f"Longitud de onda inicial: {self.start_wavelength.get()} nm"
            ws['A13'] = f"Longitud de onda final: {self.end_wavelength.get()} nm"
            ws['A14'] = f"Paso: {self.step.get()} nm"
            ws['A15'] = f"Lecturas por punto: {self.readings_per_wavelength.get()}"
            ws['A16'] = f"Tiempo de espera: {self.wait_time.get()} s"
            
            wb.save(filename)
            self.status_var.set(f"Datos guardados en {filename}")
            
        except Exception as e:
            messagebox.showerror("Error al guardar", f"No se pudo guardar el archivo Excel:\n{e}")

    def on_closing(self):
        """Manejar el cierre de la aplicación"""
        if self.measuring:
            if messagebox.askokcancel("Salir", "La medición está en progreso. ¿Está seguro de que desea salir?"):
                self.measuring = False
                # Dar tiempo para que el hilo se detenga
                self.after(500, self.destroy)
        else:
            self.destroy()

if __name__ == "__main__":
    app = FotoluminiscenciaApp()
    app.mainloop()
